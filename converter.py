from flask import Flask, request, send_file, render_template_string
import pandas as pd
import os
import chardet
from openpyxl import load_workbook

app = Flask(__name__)

# Paths for output files
RAW_EXCEL_PATH = "raw_output.xlsx"
SUMMARY_EXCEL_PATH = "summary_output.xlsx"
ALLOWED_EXTENSIONS = {'csv'}

# HTML Template
UPLOAD_FORM_HTML = """
<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
    <title>CSV to Excel Converter</title>
    <style>
      body {
        font-family: Arial, sans-serif;
        background-color: #f8f9fa;
        margin: 0;
        padding: 20px;
        display: flex;
        justify-content: center;
        align-items: center;
        flex-direction: column;
        height: 100vh;
        background-image: url('/static/background1.webp');
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
      }
      .container {
        background-color: #ffffff;
        padding: 30px;
        border-radius: 10px;
        box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
        text-align: center;
        width: 80%;
        overflow: auto;
        max-height: 90vh;
      }
      h1 {
        margin-bottom: 20px;
        color: green;
      }
      form {
        display: flex;
        justify-content: center;
        gap: 10px;
        flex-wrap: wrap;
      }
      input[type="file"] {
        margin-bottom: 10px;
      }
      input[type="submit"], a.download-btn {
        background-color: #007bff;
        color: #ffffff;
        border: none;
        padding: 10px 20px;
        border-radius: 5px;
        cursor: pointer;
        text-decoration: none;
        display: inline-block;
      }
      input[type="submit"]:hover, a.download-btn:hover {
        background-color: #0056b3;
      }
      img {
        max-width: 150px;
        margin-bottom: 20px;
      }
      table {
        width: 100%;
        border-collapse: collapse;
        margin-top: 20px;
      }
      table, th, td {
        border: 1px solid #ddd;
        padding: 8px;
        text-align: left;
      }
      th {
        background-color: #007bff;
        color: white;
        text-transform: uppercase;
      }
      td {
        max-width: 200px;
        word-wrap: break-word;
      }
      table tbody tr:nth-child(odd) {
        background-color: #f9f9f9;
      }
      table tbody tr:hover {
        background-color: #f1f1f1;
      }
      .scrollable {
        overflow: auto;
        max-height: 400px;
      }
    </style>
  </head>
  <body>
    <div class="container">
      <img src="/static/logo-avocarbon (1).png" alt="Logo">
      <h1>Upload CSV to Convert to Excel</h1>
      <form action="/convert" method="post" enctype="multipart/form-data">
        <input type="file" name="file" accept=".csv" required>
        <input type="submit" value="Convert">
      </form>
      {% if raw_table %}
      <h2>Raw Data Preview</h2>
      <div class="scrollable">
        <table>{{ raw_table|safe }}</table>
      </div>
      {% endif %}
      {% if summary_table %}
      <h2>Summary Data Preview</h2>
      <div class="scrollable">
        <table>{{ summary_table|safe }}</table>
      </div>
      {% endif %}
      {% if show_download %}
      <a href="/download/raw" class="btn download-btn">Download Raw Excel</a>
      <a href="/download/summary" class="btn download-btn">Download Summary Excel</a>
      {% endif %}
    </div>
  </body>
</html>
"""

# Helper Functions
def allowed_file(filename):
    """Check if the uploaded file is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def detect_encoding(file):
    """Detect the encoding of the file."""
    raw_data = file.read()
    result = chardet.detect(raw_data)
    file.seek(0)  # Reset file pointer
    return result['encoding']

def adjust_column_width(excel_path):
    """Adjust column widths to fit content in the Excel file."""
    wb = load_workbook(excel_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get column letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Add padding
            ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(excel_path)

def process_csv(file, original_filename):
    """Read, process, and summarize the CSV file."""
    # Detect file encoding
    encoding = detect_encoding(file)
    df = pd.read_csv(file, delimiter=';', encoding=encoding, on_bad_lines='skip')

    # Convert date columns if they exist
    if 'LIVRFINLU' in df.columns:
        df['LIVRFINLU'] = pd.to_datetime(df['LIVRFINLU'], format='%d/%m/%Y', errors='coerce')
    if 'Date debut Validité' in df.columns:
        df['Date debut Validité'] = pd.to_datetime(df['Date debut Validité'], format='%Y%m%d', errors='coerce')

    # Create a summary by 'Ref' if available
    if 'Ref' in df.columns and 'Quantité' in df.columns:
        summary = df.groupby('Ref')['Quantité'].sum().reset_index()
    else:
        summary = pd.DataFrame()

    # Generate output file names based on the original filename
    base_name = os.path.splitext(original_filename)[0]  # Strip the file extension
    raw_excel_path = f"{base_name}_raw.xlsx"
    summary_excel_path = f"{base_name}_summary.xlsx"

    # Save raw and summary data to Excel
    df.to_excel(raw_excel_path, index=False)
    summary.to_excel(summary_excel_path, index=False)

    # Adjust column widths for both Excel files
    adjust_column_width(raw_excel_path)
    adjust_column_width(summary_excel_path)

    return df, summary, raw_excel_path, summary_excel_path
@app.route('/')
def index():
    """Render the upload page."""
    return render_template_string(UPLOAD_FORM_HTML, raw_table=None, summary_table=None, show_download=False)

@app.route('/convert', methods=['POST'])
def convert():
    """Handle file upload, processing, and conversion."""
    if 'file' not in request.files:
        return "No file part", 400

    file = request.files['file']
    if file.filename == '' or not allowed_file(file.filename):
        return "Invalid file type. Please upload a CSV file.", 400

    try:
        # Process the CSV file with the original filename
        raw_data, summary_data, raw_path, summary_path = process_csv(file, file.filename)

        # Convert DataFrames to HTML for preview
        raw_table_html = raw_data.head(20).to_html(classes='table', index=False)
        summary_table_html = summary_data.to_html(classes='table', index=False) if not summary_data.empty else None

        # Save file paths to global variables for download
        global RAW_EXCEL_PATH, SUMMARY_EXCEL_PATH
        RAW_EXCEL_PATH = raw_path
        SUMMARY_EXCEL_PATH = summary_path

        return render_template_string(
            UPLOAD_FORM_HTML,
            raw_table=raw_table_html,
            summary_table=summary_table_html,
            show_download=True
        )
    except Exception as e:
        return f"Error processing the file: {e}", 500


@app.route('/download/<file_type>')
def download(file_type):
    """Handle file downloads."""
    if file_type == 'raw' and os.path.exists(RAW_EXCEL_PATH):
        return send_file(RAW_EXCEL_PATH, as_attachment=True)
    elif file_type == 'summary' and os.path.exists(SUMMARY_EXCEL_PATH):
        return send_file(SUMMARY_EXCEL_PATH, as_attachment=True)
    else:
        return "File not found", 404

if __name__ == '__main__':
    app.run(debug=True, port=5000)
